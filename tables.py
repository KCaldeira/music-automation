"""Load and validate probability/volume tables from .xlsx files."""

from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


@dataclass
class Tables:
    note_probability_table: np.ndarray      # (divisions_per_cycle,)
    rest_length_table: np.ndarray            # (divisions_per_cycle, divisions_per_bar-1)
    note_length_table: np.ndarray            # (divisions_per_cycle, divisions_per_bar-1)
    pitch_probability_table: np.ndarray      # (12,)
    interval_probability_table: np.ndarray   # (25,)
    location_volume: np.ndarray              # (divisions_per_cycle,)
    pitch_volume: np.ndarray                 # (12,)


def _load_1d(path: Path) -> np.ndarray:
    """Load a 1D table: column B (index 1) as a float array, skipping header."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    values = []
    for i, row in enumerate(ws.iter_rows(min_col=2, max_col=2, values_only=True)):
        if i == 0:
            continue  # skip header
        values.append(float(row[0]))
    wb.close()
    return np.array(values)


def _load_2d(path: Path) -> np.ndarray:
    """Load a 2D table: columns B onward as a float 2D array, skipping header."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(min_col=2, values_only=True)):
        if i == 0:
            continue  # skip header
        rows.append([float(v) if v is not None else 0.0 for v in row])
    wb.close()
    return np.array(rows)


def _normalize_rows(arr: np.ndarray) -> np.ndarray:
    """Normalize each row to sum to 1. Rows that sum to 0 become uniform."""
    sums = arr.sum(axis=1, keepdims=True)
    sums[sums == 0] = 1.0
    return arr / sums


def _normalize_1d(arr: np.ndarray) -> np.ndarray:
    """Normalize a 1D array to sum to 1."""
    s = arr.sum()
    if s == 0:
        return np.ones_like(arr) / len(arr)
    return arr / s


def load_tables(table_dir: str, divisions_per_cycle: int, divisions_per_bar: int) -> Tables:
    """Load all tables from table_dir, validate shapes, normalize probabilities."""
    d = Path(table_dir)

    note_prob = _load_1d(d / "note_probability_table.xlsx")
    pitch_prob = _load_1d(d / "pitch_probability_table.xlsx")
    interval_prob = _load_1d(d / "interval_probability_table.xlsx")
    location_vol = _load_1d(d / "location_volume.xlsx")
    pitch_vol = _load_1d(d / "pitch_volume.xlsx")
    rest_len = _load_2d(d / "rest_length_table.xlsx")
    note_len = _load_2d(d / "note_length_table.xlsx")

    # Validate shapes
    max_len = divisions_per_bar - 1
    assert note_prob.shape == (divisions_per_cycle,), \
        f"note_probability_table: expected ({divisions_per_cycle},), got {note_prob.shape}"
    assert pitch_prob.shape == (12,), \
        f"pitch_probability_table: expected (12,), got {pitch_prob.shape}"
    assert interval_prob.shape == (25,), \
        f"interval_probability_table: expected (25,), got {interval_prob.shape}"
    assert location_vol.shape == (divisions_per_cycle,), \
        f"location_volume: expected ({divisions_per_cycle},), got {location_vol.shape}"
    assert pitch_vol.shape == (12,), \
        f"pitch_volume: expected (12,), got {pitch_vol.shape}"
    assert rest_len.shape == (divisions_per_cycle, max_len), \
        f"rest_length_table: expected ({divisions_per_cycle}, {max_len}), got {rest_len.shape}"
    assert note_len.shape == (divisions_per_cycle, max_len), \
        f"note_length_table: expected ({divisions_per_cycle}, {max_len}), got {note_len.shape}"

    # Normalize probability tables
    interval_prob = _normalize_1d(interval_prob)
    pitch_prob = _normalize_1d(pitch_prob)
    rest_len = _normalize_rows(rest_len)
    note_len = _normalize_rows(note_len)

    return Tables(
        note_probability_table=note_prob,
        rest_length_table=rest_len,
        note_length_table=note_len,
        pitch_probability_table=pitch_prob,
        interval_probability_table=interval_prob,
        location_volume=location_vol,
        pitch_volume=pitch_vol,
    )
